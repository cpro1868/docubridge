from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook

from docubridge.core.xlsx_ingest import parse_xlsx_file


def _write_sample_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Summary"
    sheet1.append(["Name", "Value"])
    sheet1.append(["Alpha", 1])
    sheet1.append(["Beta", 2])

    sheet2 = workbook.create_sheet("Details")
    sheet2.append(["Key", "Text"])
    sheet2.append(["Note", "Hello"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def test_parse_xlsx_file_extracts_multiple_sheets_as_markdown(tmp_path: Path) -> None:
    input_path = tmp_path / "sample.xlsx"
    _write_sample_xlsx(input_path)

    content = parse_xlsx_file(input_path)

    assert "## Summary" in content
    assert "| Name | Value |" in content
    assert "| Alpha | 1 |" in content
    assert "## Details" in content
    assert "| Key | Text |" in content
    assert "| Note | Hello |" in content


def test_parse_xlsx_file_formats_dates_and_skips_empty_rows(tmp_path: Path) -> None:
    input_path = tmp_path / "dates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dates"
    sheet.append(["When", "Label"])
    sheet.append([date(2026, 4, 9), "Launch"])
    sheet.append([None, None])
    sheet.append([date(2026, 4, 10), "Follow-up"])
    workbook.save(input_path)

    content = parse_xlsx_file(input_path)

    assert "## Dates" in content
    assert "| 2026-04-09 | Launch |" in content
    assert "| 2026-04-10 | Follow-up |" in content
    assert "|  |  |" not in content


def test_parse_xlsx_file_truncates_long_cell_values(tmp_path: Path) -> None:
    input_path = tmp_path / "long-text.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LongText"
    sheet.append(["Key", "Text"])
    sheet.append(["A", "x" * 120])
    workbook.save(input_path)

    content = parse_xlsx_file(input_path)

    assert "## LongText" in content
    assert "...(truncated)" in content
    assert "| A | " in content
    assert "x" * 120 not in content
