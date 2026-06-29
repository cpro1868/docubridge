from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

MAX_CELL_TEXT_LENGTH = 100


def _format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time() else value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value)
    if len(text) > MAX_CELL_TEXT_LENGTH:
        return f"{text[:MAX_CELL_TEXT_LENGTH]}...(truncated)"
    return text


def _sheet_to_markdown(sheet) -> list[str]:
    rows = [
        [_format_cell_value(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    rows = [row for row in rows if any(cell != "" for cell in row)]
    if not rows:
        return [f"## {sheet.title}", ""]

    header = rows[0]
    lines = [f"## {sheet.title}", ""]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows[1:]:
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(padded[: len(header)]) + " |")
    lines.append("")
    return lines


def parse_xlsx_file(path: Path) -> str:
    workbook = load_workbook(path, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.extend(_sheet_to_markdown(sheet))

    while lines and lines[-1] == "":
        lines.pop()
    return "\n".join(lines) + ("\n" if lines else "")
